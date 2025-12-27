"""Sync WhatsApp subscribers from Supabase to the daily Excel file used by
WhatsAppExcelMonitorElevenLabsV2/scripts/autoSender.js.

Goal
----
Append new active subscribers to a given DDMMYYYY.xlsx file, copying the
non-user columns from a template row.

Excel expected columns (from repo examples)
-----------------------------------------
- NOMBRES
- APELLIDO_PATERNO
- APELLIDO_MATERNO
- CELULAR
- MAIL
- CORREO
- SMS
- WHATSAPP
- TEXTO_MENSAJE

Important
---------
- We DO NOT generate the daily image here. The sending script attaches an image
  if it exists with the same name DDMMYYYY.(png/jpg/jpeg) in the same folder.
  This sync keeps that workflow intact.
"""

from __future__ import annotations

import os
import re
import sys
import json
import argparse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

import requests
from openpyxl import load_workbook


def load_dotenv_if_present(path: Path) -> None:
    """Minimal .env loader (avoids requiring python-dotenv)."""

    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip("\"").strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


EXCEL_COLUMNS = [
    "NOMBRES",
    "APELLIDO_PATERNO",
    "APELLIDO_MATERNO",
    "CELULAR",
    "MAIL",
    "CORREO",
    "SMS",
    "WHATSAPP",
    "TEXTO_MENSAJE",
]


def normalize_phone(raw: str) -> str:
    """Keep only digits."""

    return re.sub(r"\D+", "", raw or "")


def ddmmyyyy_today() -> str:
    return datetime.now().strftime("%d%m%Y")


@dataclass
class Subscriber:
    first_name: str
    last_name: str
    phone: str
    email: str
    is_active: bool


def fetch_active_subscribers_from_supabase(
    *,
    supabase_url: str,
    service_role_key: str,
) -> List[Subscriber]:
    """Fetch all active subscribers from Supabase using the REST endpoint."""

    url = supabase_url.rstrip("/") + "/rest/v1/whatsapp_subscriptions"
    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Accept": "application/json",
    }
    params = {
        "select": "first_name,last_name,phone,email,is_active",
        "is_active": "eq.true",
        "order": "created_at.asc",
    }

    resp = requests.get(url, headers=headers, params=params, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    out: List[Subscriber] = []
    for row in data:
        out.append(
            Subscriber(
                first_name=(row.get("first_name") or "").strip(),
                last_name=(row.get("last_name") or "").strip(),
                phone=normalize_phone(str(row.get("phone") or "")),
                email=(row.get("email") or "").strip(),
                is_active=bool(row.get("is_active") is True),
            )
        )
    return out


def read_excel_headers(ws) -> List[str]:
    headers = [c.value for c in ws[1]]
    return [str(h).strip() if h is not None else "" for h in headers]


def ensure_expected_headers(headers: List[str]) -> None:
    missing = [h for h in EXCEL_COLUMNS if h not in headers]
    if missing:
        raise RuntimeError(
            "Excel no tiene las columnas esperadas. Faltan: " + ", ".join(missing)
        )


def get_existing_phones(ws, headers: List[str]) -> Set[str]:
    phone_idx = headers.index("CELULAR") + 1
    existing: Set[str] = set()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=phone_idx).value
        if val is None:
            continue
        digits = normalize_phone(str(val))
        if digits:
            existing.add(digits)
    return existing


def get_template_row_values(ws, headers: List[str], template_row: int) -> Dict[str, object]:
    """Read template row into a dict by header name."""
    values: Dict[str, object] = {}
    for h in headers:
        if not h:
            continue
        col = headers.index(h) + 1
        values[h] = ws.cell(row=template_row, column=col).value
    return values


def append_subscriber_row(
    *,
    ws,
    headers: List[str],
    template_values: Dict[str, object],
    subscriber: Subscriber,
    message_text: str,
    default_sms: int,
) -> int:
    next_row = ws.max_row + 1

    # start from template
    for h, v in template_values.items():
        if h not in headers:
            continue
        col = headers.index(h) + 1
        ws.cell(row=next_row, column=col).value = v

    # override with user fields
    def set_cell(header: str, value: object) -> None:
        col = headers.index(header) + 1
        ws.cell(row=next_row, column=col).value = value

    set_cell("NOMBRES", subscriber.first_name)
    # Split last name -> paternal/maternal if possible
    last = (subscriber.last_name or "").strip()
    parts = [p for p in re.split(r"\s+", last) if p]
    paternal = parts[0] if len(parts) >= 1 else ""
    maternal = parts[1] if len(parts) >= 2 else ""
    set_cell("APELLIDO_PATERNO", paternal)
    set_cell("APELLIDO_MATERNO", maternal)
    set_cell("CELULAR", subscriber.phone)
    set_cell("MAIL", subscriber.email)
    # Keep CORREO/WHATSAPP compatible with autoSender expectations
    set_cell("CORREO", template_values.get("CORREO", "NO") or "NO")
    set_cell("WHATSAPP", template_values.get("WHATSAPP", "SI") or "SI")
    set_cell("SMS", default_sms)
    set_cell("TEXTO_MENSAJE", message_text)

    return next_row


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sincroniza suscriptores WhatsApp (Supabase) -> Excel diario"
    )
    parser.add_argument(
        "--date",
        default=ddmmyyyy_today(),
        help="Fecha del Excel en formato DDMMYYYY (por defecto hoy)",
    )
    parser.add_argument(
        "--excel-folder",
        default=str(Path("WhatsAppExcelMonitorElevenLabsV2/scripts/excel")),
        help="Carpeta donde están los Excel diarios",
    )
    parser.add_argument(
        "--template-row",
        type=int,
        default=2,
        help="Fila usada como plantilla (por defecto 2: primera fila con datos)",
    )
    parser.add_argument(
        "--default-sms",
        type=int,
        default=0,
        help="Valor SMS para filas nuevas (0=pending, 1=sent)",
    )
    parser.add_argument(
        "--message",
        default=None,
        help="Texto del mensaje a colocar en TEXTO_MENSAJE. Si no se pasa, se toma desde la fila plantilla.",
    )
    parser.add_argument(
        "--only-whatsapp",
        action="store_true",
        help="Solo agrega suscriptores marcados para WhatsApp (WHATSAPP=SI) en el Excel.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="No escribe el Excel, solo muestra qué haría",
    )
    args = parser.parse_args()

    # Load .env if present so running the script manually works out of the box.
    load_dotenv_if_present(Path(__file__).resolve().parent.parent / ".env")

    # Load env
    supabase_url = os.getenv("SUPABASE_URL") or os.getenv("VITE_SUPABASE_URL") or ""
    service_role = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or ""
    if not supabase_url or not service_role:
        print(
            "Faltan variables de entorno SUPABASE_URL y/o SUPABASE_SERVICE_ROLE_KEY (revisa .env)",
            file=sys.stderr,
        )
        return 2

    date = str(args.date).strip()
    if not re.fullmatch(r"\d{8}", date):
        print("--date debe ser DDMMYYYY", file=sys.stderr)
        return 2

    excel_folder = Path(args.excel_folder)
    excel_path = excel_folder / f"{date}.xlsx"
    excel_path_audio = excel_folder / f"{date}_A.xlsx"
    # Prefer _A if exists (same as autoSender)
    if excel_path_audio.exists():
        excel_path = excel_path_audio

    if not excel_path.exists():
        print(f"No existe el Excel: {excel_path}", file=sys.stderr)
        return 2

    subs = fetch_active_subscribers_from_supabase(
        supabase_url=supabase_url,
        service_role_key=service_role,
    )

    wb = load_workbook(excel_path)
    ws = wb.active
    headers = read_excel_headers(ws)
    ensure_expected_headers(headers)

    existing_phones = get_existing_phones(ws, headers)
    template_values = get_template_row_values(ws, headers, args.template_row)
    template_message = template_values.get("TEXTO_MENSAJE", "") or ""
    message_text = args.message if args.message is not None else str(template_message)

    # If requested, only sync to Excel when the template itself indicates it's a WhatsApp sheet.
    if args.only_whatsapp:
        template_whatsapp = str(template_values.get("WHATSAPP", "")).strip().upper()
        if template_whatsapp not in {"SI", "SÍ"}:
            print(
                json.dumps(
                    {
                        "excel": str(excel_path),
                        "skipped": True,
                        "reason": f"Plantilla WHATSAPP={template_whatsapp!r} (se esperaba 'SI')",
                    },
                    ensure_ascii=False,
                    indent=2,
                )
            )
            return 0

    to_add: List[Subscriber] = []
    for s in subs:
        if not s.phone:
            continue
        if s.phone in existing_phones:
            continue
        to_add.append(s)

    if not to_add:
        print(
            json.dumps(
                {
                    "excel": str(excel_path),
                    "active_subscribers": len(subs),
                    "existing_in_excel": len(existing_phones),
                    "added": 0,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    # Append rows
    for s in to_add:
        append_subscriber_row(
            ws=ws,
            headers=headers,
            template_values=template_values,
            subscriber=s,
            message_text=message_text,
            default_sms=args.default_sms,
        )

    result = {
        "excel": str(excel_path),
        "active_subscribers": len(subs),
        "existing_in_excel": len(existing_phones),
        "added": len(to_add),
        "note": "La imagen se envía si existe DDMMYYYY.(png/jpg/jpeg) en la misma carpeta (autoSender.js).",
    }

    if args.dry_run:
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    wb.save(excel_path)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
